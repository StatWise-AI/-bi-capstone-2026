"""
Builds the Power BI-ready Excel workbook from the agent's current data:
a star-schema-friendly Fact table + Date dimension, plus a backtest results
sheet and a methodology/assumptions sheet.

Callable directly from the Streamlit app (build_export(agent, path)) so the
"Refresh Power BI export" button in app.py regenerates the exact same file
Power BI is pointed at — no separate manual export step.

Note: this still is NOT a live link. Power BI has to click "Refresh" itself
after this file is regenerated; see README.md for the two-click workflow.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT_PATH = os.path.join(APP_DIR, "EU_Diesel_PowerBI_Data_Model.xlsx")

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F3864")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        content_max = 0
        for v in df[col].tolist():
            l = len(str(v)) if v is not None else 0
            if l > content_max:
                content_max = l
        width = max(12, min(38, max(content_max, len(str(col))) + 4))
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


def _number_format_for(col):
    if col in ("PriceWithTax_EUR_per_1000L", "PriceWoTax_EUR_per_1000L", "TaxAmount_EUR_per_1000L",
               "PriceWithTax_Lower80", "PriceWithTax_Upper80", "MeanMAE", "MeanRMSE"):
        return "#,##0.0"
    if col == "TaxSharePct":
        return "0.0%"
    return None


def write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header_row(ws, start_row, len(df.columns))
    formats = {col: _number_format_for(col) for col in df.columns}
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            val = row[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if formats[col] and val is not None:
                cell.number_format = formats[col]
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)


def build_dim_date(combined):
    dates = pd.to_datetime(combined["date"])
    dim = pd.DataFrame({"date": dates})
    dim["year"] = dim["date"].dt.year
    dim["quarter"] = "Q" + dim["date"].dt.quarter.astype(str)
    dim["month"] = dim["date"].dt.month
    dim["month_name"] = dim["date"].dt.strftime("%B")
    dim["iso_week"] = dim["date"].dt.isocalendar().week
    dim["is_forecast_period"] = combined["is_forecast"].values
    dim = dim.drop_duplicates(subset="date").sort_values("date").reset_index(drop=True)
    return dim


def build_export(agent, output_path=DEFAULT_OUT_PATH):
    """Regenerate the Power BI workbook from the agent's current state.
    Returns the output path. Call this from the Streamlit refresh button."""
    combined = agent.get_combined()
    backtest = agent.backtest_summary()
    meta = agent.get_meta()

    fact = combined.rename(columns={
        "price_with_tax_eur_per_1000l": "PriceWithTax_EUR_per_1000L",
        "price_wo_tax_eur_per_1000l": "PriceWoTax_EUR_per_1000L",
        "tax_amount_eur_per_1000l": "TaxAmount_EUR_per_1000L",
        "tax_share_of_price": "TaxSharePct",
        "is_forecast": "IsForecast",
        "forecast_method": "ForecastMethod",
        "price_with_tax_lower80": "PriceWithTax_Lower80",
        "price_with_tax_upper80": "PriceWithTax_Upper80",
    })
    fact = fact.rename(columns={"date": "Date"})
    fact["Date"] = pd.to_datetime(fact["Date"]).dt.date
    # Newest-first: present -> 2005, per user preference for all exports.
    fact = fact.sort_values("Date", ascending=False).reset_index(drop=True)

    dim_date = build_dim_date(combined)
    dim_date = dim_date.rename(columns={
        "date": "Date", "year": "Year", "quarter": "Quarter", "month": "Month",
        "month_name": "MonthName", "iso_week": "ISOWeek", "is_forecast_period": "IsForecastPeriod",
    })
    dim_date["Date"] = pd.to_datetime(dim_date["Date"]).dt.date
    dim_date = dim_date.sort_values("Date", ascending=False).reset_index(drop=True)

    backtest_out = backtest.rename(columns={
        "method": "Method", "mean_mae": "MeanMAE", "mean_rmse": "MeanRMSE", "n_folds": "NumFolds",
    })
    backtest_out["Method"] = backtest_out["Method"].str.upper()
    backtest_out[["MeanMAE", "MeanRMSE"]] = backtest_out[["MeanMAE", "MeanRMSE"]].round(2)

    wb = Workbook()

    ws = wb.active
    ws.title = "Fact_DieselPrice"
    write_df(ws, fact)
    autosize(ws, fact)

    ws2 = wb.create_sheet("Dim_Date")
    write_df(ws2, dim_date)
    autosize(ws2, dim_date)

    ws3 = wb.create_sheet("Backtest_Results")
    ws3.cell(row=1, column=1,
             value=f"Model comparison — rolling-origin backtest ({meta.get('n_backtest_origins', 'multiple')} fold origins)").font = TITLE_FONT
    write_df(ws3, backtest_out, start_row=3)
    autosize(ws3, backtest_out)
    note_row = 3 + len(backtest_out) + 2
    ws3.cell(row=note_row, column=1,
             value="Lower MAE/RMSE = better. Backtest run on the pre-tax (market) price series only.").font = NOTE_FONT

    ws4 = wb.create_sheet("Methodology_Notes")
    ws4.cell(row=1, column=1, value="Methodology and assumptions").font = TITLE_FONT
    scope_line = (f"Scope: {meta.get('country', 'EU aggregate')} / {meta.get('product', 'diesel')}"
                  if meta.get("country") else "Scope: EU aggregate / diesel")
    notes = [
        "",
        "Source: European Commission Weekly Oil Bulletin (DG Energy).",
        scope_line,
        f"Coverage: weekly, through {meta['last_history_date']}.",
        "",
        f"Forecast horizon: {meta['horizon_weeks']} weeks ahead from {meta['last_history_date']}.",
        f"Forecast model (selected via AIC grid search): {meta['model_name']}, "
        "fit on the PRE-TAX (market) price series.",
        "",
        "Why forecast the pre-tax series instead of the retail price directly: the pre-tax price moves "
        "with crude/refining markets and is what time-series methods are suited to. Tax (VAT + excise + "
        "other indirect taxes) is policy-driven and changes in discrete steps, not a trend.",
        "",
        f"With-tax reconstruction: forecast(with-tax) = forecast(pre-tax) + latest observed tax amount "
        f"(EUR {meta['latest_tax_amount']:.1f} per 1000L, held flat as a documented assumption).",
        "",
        "Caveat: any scheduled excise duty or VAT change within the forecast window would shift the "
        "with-tax figure without the underlying market forecast changing. For country/product views, "
        "Brent crude and (where applicable) the exchange rate are also held flat at their latest observed "
        "level over the forecast horizon -- see the model comparison document for the full methodology.",
        "",
        "This file was regenerated from the Streamlit app's 'Refresh Power BI export' button, using "
        "whichever region/product was selected at the time. In Power BI, use Home > Refresh after this "
        "file updates to pull in the new values.",
    ]
    for i, line in enumerate(notes):
        cell = ws4.cell(row=2 + i, column=1, value=line)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.column_dimensions["A"].width = 110

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    from diesel_agent import DieselForecastAgent
    agent = DieselForecastAgent()
    path = build_export(agent)
    print(f"Saved: {path}")
