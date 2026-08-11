"""
build_full_powerbi_export.py
==============================
Builds ONE Power BI-ready workbook covering ALL 52 country/product
combinations -- unlike the Streamlit app's "Download Power BI workbook"
button, which only exports whichever single combination is currently
selected in the sidebar.

Use this when you want the same all-country, all-product dashboard
experience in Power BI that you get in Streamlit (with Country/Product
slicers), rather than one static report per combination.

Sheets produced:
  - Fact_DieselPrice   -- all 52 combinations, one row per (country, product,
                           date), with the model actually used per row
  - Dim_Date           -- standard date dimension for time intelligence
  - Backtest_Results   -- every method tested per combination (naive, ARIMA,
                           SARIMAX, and GBM where it was tested), so the
                           "which model, and why" story is in the data too
  - Methodology_Notes  -- assumptions and caveats, same as the single-combo
                           export

Column names match measures_tmdl_view_script.tmdl and the generated TMDL
files exactly, so the DAX measures from Part 1 of the build guide work
against this file without modification.
"""
import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "Barrl_PowerBI_Full_Dashboard_Data.xlsx"

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F3864")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COUNTRY_LABELS = {
    "DE": "Germany", "FR": "France", "IT": "Italy", "ES": "Spain",
    "NL": "Netherlands", "BE": "Belgium", "PL": "Poland", "AT": "Austria",
    "CZ": "Czechia", "PT": "Portugal",
}
PRODUCT_LABELS = {
    "euro95": "Petrol (Euro-95)", "diesel": "Diesel", "heating_oil": "Heating oil",
    "fuel_oil_1": "Fuel oil (\u22641% sulphur)", "fuel_oil_2": "Fuel oil (>1% sulphur)", "lpg": "LPG",
}


def style_header_row(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, df, start_col=1):
    for i, col in enumerate(df.columns):
        content_max = 0
        for v in df[col].tolist():
            l = len(str(v)) if v is not None else 0
            if l > content_max:
                content_max = l
        width = max(12, min(38, max(content_max, len(str(col))) + 4))
        ws.column_dimensions[get_column_letter(start_col + i)].width = width


def _number_format_for(col):
    if col == "Date":
        return "yyyy-mm-dd"
    if col in ("PriceWithTax_EUR_per_1000L", "PriceWoTax_EUR_per_1000L", "TaxAmount_EUR_per_1000L",
               "PriceWithTax_Lower80", "PriceWithTax_Upper80", "MeanMAE", "MeanRMSE"):
        return "#,##0.0"
    if col == "TaxSharePct":
        return "0.0%"
    return None


def write_df(ws, df, start_row=1, start_col=1):
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col)
    style_header_row(ws, start_row, len(df.columns))
    formats = {col: _number_format_for(col) for col in df.columns}
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            val = row[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if formats[col] and val is not None:
                cell.number_format = formats[col]
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col + 2)


def build_dim_date(all_dates):
    # Power BI's "Mark as date table" requires a genuinely continuous
    # calendar -- one row per day, no gaps -- even though the underlying
    # fact data is weekly. Most of these dates simply won't have a matching
    # fact row, which is normal and expected for a date dimension.
    min_date = pd.to_datetime(all_dates.min())
    max_date = pd.to_datetime(all_dates.max())
    dim = pd.DataFrame({"Date": pd.date_range(min_date, max_date, freq="D")})
    dim["Year"] = dim["Date"].dt.year
    dim["Quarter"] = "Q" + dim["Date"].dt.quarter.astype(str)
    dim["Month"] = dim["Date"].dt.month
    dim["MonthName"] = dim["Date"].dt.strftime("%B")
    dim["ISOWeek"] = dim["Date"].dt.isocalendar().week
    dim["Date"] = dim["Date"].dt.date
    return dim


def main():
    combined = pd.read_csv("country_product_history_and_forecast.csv", parse_dates=["date"])
    meta = json.load(open("country_product_meta.json"))
    model_lookup = {(m["country"], m["product"]): m for m in meta}

    combined["tax_amount"] = combined["price_with_tax"] - combined["price_wo_tax"]
    combined["tax_share"] = combined["tax_amount"] / combined["price_with_tax"]

    def method_label(row):
        m = model_lookup.get((row["country"], row["product"]), {})
        if not row["is_forecast"]:
            return None
        if m.get("model_used") == "gbm":
            return "LightGBM (direct multi-horizon)"
        order = m.get("sarimax_order")
        return f"SARIMAX{tuple(order)}" if order else "SARIMAX"

    combined["forecast_method"] = combined.apply(method_label, axis=1)

    fact = combined.rename(columns={
        "country": "Country", "product": "Product", "date": "Date",
        "price_with_tax": "PriceWithTax_EUR_per_1000L",
        "price_wo_tax": "PriceWoTax_EUR_per_1000L",
        "tax_amount": "TaxAmount_EUR_per_1000L",
        "tax_share": "TaxSharePct",
        "is_forecast": "IsForecast",
        "forecast_method": "ForecastMethod",
        "price_with_tax_lower80": "PriceWithTax_Lower80",
        "price_with_tax_upper80": "PriceWithTax_Upper80",
    })
    fact["Country"] = fact["Country"].map(lambda c: COUNTRY_LABELS.get(c, c))
    fact["Product"] = fact["Product"].map(lambda p: PRODUCT_LABELS.get(p, p))
    fact["Date"] = pd.to_datetime(fact["Date"]).dt.date
    fact = fact.sort_values("Date", ascending=False).reset_index(drop=True)
    fact = fact[["Country", "Product", "Date", "PriceWithTax_EUR_per_1000L", "PriceWoTax_EUR_per_1000L",
                 "TaxAmount_EUR_per_1000L", "TaxSharePct", "IsForecast", "ForecastMethod",
                 "PriceWithTax_Lower80", "PriceWithTax_Upper80"]]

    dim_date = build_dim_date(combined["date"])
    dim_date = dim_date.sort_values("Date", ascending=False).reset_index(drop=True)

    bt_rows = []
    for m in meta:
        bt = m.get("backtest", {})
        for method_key, label in [("naive", "Naive"), ("arima_noexog", "ARIMA (no exog)"),
                                   ("sarimax_exog", "SARIMAX (+exog)"), ("gbm", "GBM")]:
            if method_key in bt:
                bt_rows.append({
                    "Country": COUNTRY_LABELS.get(m["country"], m["country"]),
                    "Product": PRODUCT_LABELS.get(m["product"], m["product"]),
                    "Method": label,
                    "MeanMAE": round(bt[method_key]["mean_mae"], 2),
                    "MeanRMSE": round(bt[method_key]["mean_rmse"], 2),
                    "Folds": bt[method_key]["n_folds"],
                    "IsProductionModel": (
                        (label == "GBM" and m.get("model_used") == "gbm") or
                        (label == "SARIMAX (+exog)" and m.get("model_used") != "gbm")
                    ),
                })
    backtest_df = pd.DataFrame(bt_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fact_DieselPrice"
    write_df(ws, fact)
    autosize(ws, fact)

    ws2 = wb.create_sheet("Dim_Date")
    write_df(ws2, dim_date)
    autosize(ws2, dim_date)

    ws3 = wb.create_sheet("Backtest_Results")
    write_df(ws3, backtest_df, start_row=1)
    autosize(ws3, backtest_df)
    note_row = 1 + len(backtest_df) + 2
    ws3.cell(row=note_row, column=1,
             value="Note: IsProductionModel = TRUE marks whichever method is actually used for that combination's live forecast. "
                   "(This note sits below the data table on purpose -- Power BI's importer expects headers in row 1.)").font = NOTE_FONT

    ws4 = wb.create_sheet("Methodology_Notes")
    ws4.cell(row=1, column=1, value="Methodology and assumptions").font = TITLE_FONT
    notes = [
        "",
        "Source: European Commission Weekly Oil Bulletin (DG Energy), all 10 countries in scope, all 6 products where data allows.",
        "52 of 60 possible (country, product) combinations included -- see completeness_report.csv for exclusions.",
        "",
        "Forecast model varies by combination: SARIMAX with exogenous regressors (Brent crude, FX for non-Eurozone "
        "countries, seasonality) for most combinations; LightGBM (gradient-boosted trees) for diesel (all 10 countries) "
        "and heating oil (8 of 10 -- Germany and the Netherlands specifically keep SARIMAX). See ForecastMethod in "
        "Fact_DieselPrice for which model produced each forecast row, and Backtest_Results for the evidence behind "
        "each choice -- see Model_Comparison_Report.docx, Section 8, for the full reasoning.",
        "",
        "Tax reconstruction: forecast(with-tax) = forecast(pre-tax) + latest observed tax amount, held flat. This is "
        "a documented simplifying assumption, not a prediction that tax policy won't change.",
        "",
        "This file covers ALL combinations in one workbook, unlike the Streamlit app's single-combination Power BI "
        "export button. Use this file for a Country/Product-sliceable dashboard; use the app's own export for a "
        "one-off single-combination snapshot.",
    ]
    for i, line in enumerate(notes):
        cell = ws4.cell(row=2 + i, column=1, value=line)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.column_dimensions["A"].width = 110

    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")
    print(f"Fact_DieselPrice: {len(fact)} rows, {fact['Country'].nunique()} countries, {fact['Product'].nunique()} products")
    print(f"Dim_Date: {len(dim_date)} rows")
    print(f"Backtest_Results: {len(backtest_df)} rows")


if __name__ == "__main__":
    main()
