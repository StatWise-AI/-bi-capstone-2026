"""
Extraction pipeline for the country-level, multi-product upgrade.

Builds one long-format table: one row per (country, product, date), with:
  - price_with_tax, price_wo_tax  (target + market/pre-tax component)
  - exchange_rate  (to national currency; 1.0 constant for Eurozone countries)
  - brent_usd_per_bbl  (monthly, forward-filled to weekly)
  - month, quarter  (seasonal indicators)

IMPORTANT DATA-QUALITY FINDING (discovered during implementation, not assumed
up front): the workbook's "VAT" and "Excise duties" sheets, despite having a
"Since:" date column that looks like a change-history log, in practice
contain only ONE row per country — the CURRENT rate, not a historical
series. Confirmed by direct inspection: e.g. Germany's excise duty sheet has
a single row dated 2026-05-01; there is no 2005-2025 history to merge onto
our weekly panel. Treating a constant as a time-varying regressor would be
statistically meaningless (no variation to learn from) and would overstate
what the data supports.

Consequence: VAT/excise are NOT used as SARIMAX exogenous regressors here.
Instead, the realized tax burden (price_with_tax - price_wo_tax), which DOES
have full weekly history directly from the price sheets, is used for the
market/tax decomposition and for reconstructing the with-tax forecast --
exactly the same approach validated in the EU-aggregate model. The current
VAT/excise snapshot is still extracted and saved separately for reference/
documentation (see current_tax_rates.csv), since it's genuinely useful
context even though it can't be a model input.

Country/product coverage is checked for completeness; combinations with too
little price history are excluded and logged explicitly (completeness_report.csv).
"""
import openpyxl
import pandas as pd
import numpy as np
import datetime
import os

SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx")
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
BRENT_PATH = os.path.join(OUT_DIR, "brent_crude_monthly.csv")

TARGET_COUNTRIES = ['DE_', 'FR_', 'IT_', 'ES_', 'NL_', 'BE_', 'PL_', 'AT_', 'CZ_', 'PT_']
PRODUCTS = ['euro95', 'diesel', 'heating_oil', 'fuel_oil_1', 'fuel_oil_2', 'lpg']
MIN_WEEKS_REQUIRED = 150  # ~3 years; below this a country/product combo is excluded

wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)


def get_price_blocks(sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=4, max_col=226, values_only=True))
    header1, row4 = rows[0], rows[3]
    ctr_cols = [i for i, v in enumerate(header1) if v == 'CTR']
    blocks = {}
    for j, idx in enumerate(ctr_cols):
        code = row4[idx]
        if code is None:
            continue
        nxt = ctr_cols[j + 1] if j + 1 < len(ctr_cols) else 226
        blocks[code] = (idx, nxt - idx)
    return blocks


def extract_price_series(sheet_name, country_code, blocks):
    idx, width = blocks[country_code]
    has_fx = (width == 8)
    prod_offset = 2 if has_fx else 1
    ws = wb[sheet_name]
    max_col = idx + width
    out = {"date": [], "exchange_rate": []}
    for p in PRODUCTS:
        out[p] = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
        date = row[0]
        if not isinstance(date, datetime.datetime):
            continue
        out["date"].append(date.date())
        out["exchange_rate"].append(row[idx + 1] if has_fx else None)
        for k, p in enumerate(PRODUCTS):
            out[p].append(row[idx + prod_offset + k])
    df = pd.DataFrame(out)
    return df, has_fx


def extract_current_tax_rate(sheet_name, country_code):
    """VAT / Excise duties sheets: confirmed to hold only the CURRENT rate
    per country (single row), not a historical log. Returns that snapshot."""
    ws = wb[sheet_name]
    latest = None
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=8, values_only=True):
        if row[0] != country_code:
            continue
        since = row[1]
        if not isinstance(since, datetime.datetime):
            continue
        if latest is None or since > latest[0]:
            latest = (since, row[2:8])
    if latest is None:
        return None, [None] * 6
    return latest[0].date(), list(latest[1])


# ---------------------------------------------------------------------
# Brent crude: monthly -> forward-filled to weekly
# ---------------------------------------------------------------------
brent = pd.read_csv(BRENT_PATH, parse_dates=["date"]).set_index("date")["brent_usd_per_bbl"]


def brent_for_dates(weekly_dates):
    idx = pd.to_datetime(weekly_dates)
    monthly_reindexed = brent.reindex(pd.date_range(brent.index.min(), idx.max() + pd.Timedelta(days=31), freq="D")).ffill()
    return monthly_reindexed.reindex(idx, method="ffill")


# ---------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------
blocks_wt = get_price_blocks("Prices with taxes")
blocks_wo = get_price_blocks("Prices wo taxes")

all_rows = []
completeness = []
current_tax_rows = []

for country in TARGET_COUNTRIES:
    wt_df, has_fx = extract_price_series("Prices with taxes", country, blocks_wt)
    wo_df, _ = extract_price_series("Prices wo taxes", country, blocks_wo)

    vat_date, vat_vals = extract_current_tax_rate("VAT", country)
    excise_date, excise_vals = extract_current_tax_rate("Excise duties", country)
    for k, p in enumerate(PRODUCTS):
        current_tax_rows.append({
            "country": country.strip("_"), "product": p,
            "current_vat_pct": vat_vals[k], "vat_effective_since": vat_date,
            "current_excise_duty": excise_vals[k], "excise_effective_since": excise_date,
        })

    weekly_dates = wt_df["date"].tolist()
    brent_series = brent_for_dates(weekly_dates).values

    for product in PRODUCTS:
        with_tax = wt_df[product]
        wo_tax = wo_df[product]
        n_valid = with_tax.notna().sum()
        completeness.append({
            "country": country, "product": product,
            "n_weeks_available": int(n_valid),
            "first_date": wt_df.loc[with_tax.notna(), "date"].min() if n_valid else None,
            "last_date": wt_df.loc[with_tax.notna(), "date"].max() if n_valid else None,
            "included": bool(n_valid >= MIN_WEEKS_REQUIRED),
        })
        if n_valid < MIN_WEEKS_REQUIRED:
            continue

        block = pd.DataFrame({
            "country": country.strip("_"),
            "product": product,
            "date": weekly_dates,
            "price_with_tax": with_tax.values,
            "price_wo_tax": wo_tax.values,
            "exchange_rate": wt_df["exchange_rate"].values if has_fx else 1.0,
            "is_eurozone": not has_fx,
            "brent_usd_per_bbl": brent_series,
        })
        all_rows.append(block)

master = pd.concat(all_rows, ignore_index=True)
master["date"] = pd.to_datetime(master["date"])
master["tax_amount"] = master["price_with_tax"] - master["price_wo_tax"]
master["tax_share"] = master["tax_amount"] / master["price_with_tax"]
master["month"] = master["date"].dt.month
master["quarter"] = master["date"].dt.quarter
master = master.sort_values(["country", "product", "date"]).reset_index(drop=True)
master = master.dropna(subset=["price_with_tax", "price_wo_tax"])

master.to_csv(os.path.join(OUT_DIR, "master_country_product_weekly.csv"), index=False)

comp_df = pd.DataFrame(completeness)
comp_df.to_csv(os.path.join(OUT_DIR, "completeness_report.csv"), index=False)

tax_df = pd.DataFrame(current_tax_rows)
tax_df.to_csv(os.path.join(OUT_DIR, "current_tax_rates.csv"), index=False)

print("Master table shape:", master.shape)
print("NaN check on modeling columns:")
print(master[["price_with_tax", "price_wo_tax", "exchange_rate", "brent_usd_per_bbl"]].isna().sum())
print("\nIncluded combinations:", comp_df["included"].sum(), "/", len(comp_df))
print("\nExcluded combinations:")
print(comp_df[~comp_df["included"]][["country", "product", "n_weeks_available"]].to_string(index=False))
print("\nCurrent tax rate snapshot (sample):")
print(tax_df.head(8).to_string(index=False))
